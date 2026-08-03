#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ETL de VALORES_DE_ORIENTACION.xlsx — la fuente primaria del laboratorio.

Lee las tres hojas (FQ / CR / OTROS) SIN transcribir nada a mano: los grupos de
columnas se resuelven desde las celdas combinadas del propio archivo, y cada
valor sale con su celda de origen (hoja, fila, columna) para que el cotejo sea
auditable.

Produce dos archivos:

  1. database/seeders/data/valores_orientacion_oficial.json
     La matriz COMPLETA del Excel (todos los fluidos × estados × clases de
     tensión × equipos), cada límite con su origen. Es el universo del que el
     laboratorio elige qué activar — en el sistema viejo eligió un subconjunto
     y lo clavó en código; acá la elección es una fila de datos.

  2. docs/migracion/auditoria/COTEJO-EXCEL-LIMITES.md
     El cotejo contra lo SEMBRADO (spec_limits_legacy.json, extraído del código
     Ruby), en cuatro categorías: coinciden · difieren · solo en el Excel ·
     solo en el código. NO cambia ningún número: es el informe con el que el
     laboratorio decide.

Correr desde la raíz del proyecto:  python3 docs/origen-ruby/fuentes-laboratorio/etl_valores_orientacion.py
"""

import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parents[3]
EXCEL = RAIZ / 'docs/origen-ruby/fuentes-laboratorio/VALORES_DE_ORIENTACION.xlsx'
LEGADO = RAIZ / 'database/seeders/data/spec_limits_legacy.json'
SALIDA_JSON = RAIZ / 'database/seeders/data/valores_orientacion_oficial.json'
SALIDA_COTEJO = RAIZ / 'docs/migracion/auditoria/COTEJO-EXCEL-LIMITES.md'


def limpiar(v):
    if v is None:
        return ''
    return ' '.join(str(v).split())


def parsear_valor(texto, sentido):
    """'0.20 - máximo' → (<=, 0.20) · '40' con sentido mínimo → (>=, 40.0) ·
    '-' → (None, None) · texto no numérico → ('text', None)."""
    t = limpiar(texto)
    if t in ('', '-', '–'):
        return None, None, t or '-'
    # ** del acetileno del horno: el Excel remite a nota, no da número.
    if t == '**':
        return None, None, t
    m = re.search(r'-?\d+(?:[.,]\d+)?', t.replace('<', ' ').replace('≥', ' ').replace('≤', ' '))
    if m is None:
        return 'text', None, t
    numero = float(m.group(0).replace(',', '.'))
    bajo = t.lower()
    if 'mín' in bajo or 'min' in bajo:
        op = '>='
    elif 'máx' in bajo or 'max' in bajo:
        op = '<='
    elif t.strip().startswith('<'):
        op = '<'
    elif t.strip().startswith(('≥', '>')):
        op = '>='
    elif sentido == 'mínimo':
        op = '>='
    elif sentido == 'máximo':
        op = '<='
    else:
        op = None
    return op, numero, t


def rangos_combinados(ws):
    """cel (fila, col) → (fila, col) de la esquina superior izquierda de su rango."""
    mapa = {}
    for rango in ws.merged_cells.ranges:
        for fila in range(rango.min_row, rango.max_row + 1):
            for col in range(rango.min_col, rango.max_col + 1):
                mapa[(fila, col)] = (rango.min_row, rango.min_col)
    return mapa


def valor_resuelto(ws, mapa, fila, col):
    origen = mapa.get((fila, col), (fila, col))
    return limpiar(ws.cell(*origen).value)


def sin_acentos(t):
    return ''.join(c for c in unicodedata.normalize('NFD', t) if unicodedata.category(c) != 'Mn')


# ── FQ ────────────────────────────────────────────────────────────────────────

# El nombre del parámetro + su norma → el código que usa spec_limits_legacy.json
# (mismos códigos para que el cotejo sea directo). Los pares (parámetro, norma)
# que el código viejo NO tenía reciben código propio.
FQ_PARAMETROS = {
    ('ACIDEZ', 'ASTM 974'): 'acid',
    ('FACT. DE POT. 25°C', 'ASTM 924'): 'pot@25',
    ('FACT. DE POT. 25°C', 'IEC60247'): 'pot@25_iec60247',
    ('FACT. DE POT.90°C', 'IEC60247'): 'pot@90',
    ('FACT. DE POT.100°C', 'ASTM 924'): 'pot@100',
    ('RIGIDEZ', 'ASTM 1816'): 'rig_d1816',
    ('RIGIDEZ', 'ASTM 877'): 'rig_d877',
    ('RIGIDEZ', 'IEC 60156'): 'rig_iec60156',
    ('TENSION', 'ASTM 971'): 'ten',
    ('HUMEDAD', 'ASTM 1533'): 'wat',
    ('COLOR', 'ASTM1500'): 'col',
    ('VISUAL', 'ASTM 1524'): 'con',
    ('DENSIDAD (max)', 'ASTM 7777'): 'den',
}


def normalizar_estado(texto):
    t = sin_acentos(limpiar(texto)).upper()
    if 'ANTES DE ENERGIZAR' in t:
        return 'antes_de_energizar'
    if 'NUEVO EN TRAFO' in t:
        return 'nuevo_en_trafo'
    if 'EN SERVICIO' in t:
        return 'en_servicio'
    if 'TRATADO' in t:
        return 'tratado'
    if 'USADO' in t:
        return 'usado'
    if 'NUEVO' in t:
        return 'nuevo'
    return None


def normalizar_kv(texto):
    t = limpiar(texto)
    if t in ('', '-'):
        return None
    if t.upper() == 'NEUTRAL':
        return 'neutral'
    t = t.replace(' ', '').replace('KV', '').replace('kV', '')
    return t   # '≤69', '>69-<230', '≥230-<345', '≥230', '>69', '>69-230'


def extraer_fq(wb):
    ws = wb['FQ']
    mapa = rangos_combinados(ws)
    filas = []

    # El grupo (fluido+estado) sale de la fila 3 (o 4 para el conmutador) vía
    # celdas combinadas; la clase de tensión de la fila 5.
    for col in range(5, ws.max_column + 1):
        grupo3 = valor_resuelto(ws, mapa, 3, col)
        grupo4 = valor_resuelto(ws, mapa, 4, col)
        encabezado = grupo4 or grupo3
        if not encabezado:
            continue
        enc = sin_acentos(encabezado).upper()
        if 'CONMUTADOR' in enc:
            fluido = 'conmutador'
        elif 'MINERAL' in enc:
            fluido = 'mineral'
        elif 'SILICONA' in enc:
            fluido = 'silicona'
        elif 'MIDEL' in enc:
            fluido = 'midel'
        elif 'VEGETAL' in enc:
            fluido = 'vegetal'
        else:
            continue
        estado = normalizar_estado(encabezado)
        kv = normalizar_kv(ws.cell(5, col).value)

        for fila in range(6, ws.max_row + 1):
            nombre = limpiar(ws.cell(fila, 2).value)
            norma = limpiar(ws.cell(fila, 3).value)
            sentido = limpiar(ws.cell(fila, 4).value)
            if not nombre:
                continue
            clave = (nombre, norma)
            if clave not in FQ_PARAMETROS:
                raise SystemExit(f'FQ fila {fila}: parámetro sin mapa: {clave}')
            op, numero, display = parsear_valor(ws.cell(fila, col).value, sentido)
            filas.append({
                'hoja': 'FQ', 'celda': f'{ws.cell(fila, col).coordinate}',
                'fluido': fluido, 'estado': estado, 'kv': kv, 'equipo': None,
                'analito': FQ_PARAMETROS[clave], 'norma_metodo': norma,
                'operador': op, 'valor': numero, 'display': display,
            })
    return filas


# ── CR ────────────────────────────────────────────────────────────────────────

CR_GASES = ['h2', 'o2', 'n2', 'ch4', 'co', 'co2', 'c2h4', 'c2h6', 'c2h2']

CR_EQUIPOS = {
    'DISTRIBUCION <5MVA': 'distribucion',
    'POTENCIA ≥5MVA': 'potencia',
    'DE HORNO': 'horno',
    'DE CORRIENTE': 'corriente',
    'DE VOLTAJE': 'voltaje',
    'INSTRUMENTO': 'instrumento',
    'BUSHING': 'bushing',
    'CABLES': 'cables',
    'INTERRUPTOR': 'interruptor',
    'ACEITE SOYA - FR3': None,     # variante del fluido, no equipo
    'ACEITE GIRASOL': None,
}


def extraer_cr(wb):
    ws = wb['CR']
    mapa = rangos_combinados(ws)
    filas = []
    for fila in range(5, ws.max_row + 1):
        fluido_crudo = valor_resuelto(ws, mapa, fila, 1)
        tipo_crudo = limpiar(ws.cell(fila, 2).value)
        if not fluido_crudo:
            continue
        fluido = sin_acentos(fluido_crudo).lower()
        tipo_norm = sin_acentos(tipo_crudo).upper()
        equipo = CR_EQUIPOS.get(tipo_norm, None) if tipo_crudo else None
        variante = None
        if tipo_norm == 'ACEITE SOYA - FR3':
            variante = 'soya_fr3'
        elif tipo_norm == 'ACEITE GIRASOL':
            variante = 'girasol'
        elif tipo_crudo and equipo is None:
            raise SystemExit(f'CR fila {fila}: tipo sin mapa: {tipo_crudo!r}')
        for i, gas in enumerate(CR_GASES):
            op, numero, display = parsear_valor(ws.cell(fila, 3 + i).value, 'máximo')
            filas.append({
                'hoja': 'CR', 'celda': ws.cell(fila, 3 + i).coordinate,
                'fluido': fluido, 'estado': None, 'kv': None,
                'equipo': equipo, 'variante': variante,
                'analito': gas, 'norma_metodo': None,
                'operador': op, 'valor': numero, 'display': display,
            })
    return filas


# ── OTROS ─────────────────────────────────────────────────────────────────────

def extraer_otros(wb):
    ws = wb['OTROS']
    mapa = rangos_combinados(ws)
    filas = []

    # Columnas: 5 = general/mineral; 6-8 silicona nuevo/nuevo-en-trafo/usado;
    # 9-15 vegetal (nuevo, antes ≤69/69-230/≥230, usado ≤69/69-230/≥230).
    columnas = []
    for col in range(5, ws.max_column + 1):
        enc2 = valor_resuelto(ws, mapa, 2, col)
        kv = normalizar_kv(ws.cell(4, col).value)
        enc = sin_acentos(enc2).upper()
        if 'SILICONA' in enc:
            fluido, estado = 'silicona', normalizar_estado(enc2)
        elif 'VEGETAL' in enc:
            fluido, estado = 'vegetal', normalizar_estado(enc2)
        else:
            fluido, estado = 'general', None
        columnas.append((col, fluido, estado, kv))

    # Los parámetros multibanda (grado de polimerización, pasivador) apilan sus
    # bandas en filas sin nombre: heredan el último parámetro visto.
    ultimo = None
    for fila in range(5, ws.max_row + 1):
        nombre = limpiar(ws.cell(fila, 2).value)
        norma = limpiar(ws.cell(fila, 3).value)
        sentido = limpiar(ws.cell(fila, 4).value)
        if nombre:
            ultimo = (nombre, norma, sentido)
        elif ultimo is None:
            continue
        nombre_ef, norma_ef, sentido_ef = (nombre or ultimo[0], norma or ultimo[1], sentido or ultimo[2])
        for col, fluido, estado, kv in columnas:
            crudo = ws.cell(fila, col).value
            if limpiar(crudo) == '':
                continue
            op, numero, display = parsear_valor(crudo, sentido_ef)
            filas.append({
                'hoja': 'OTROS', 'celda': ws.cell(fila, col).coordinate,
                'fluido': fluido, 'estado': estado, 'kv': kv, 'equipo': None,
                'analito': None, 'parametro': nombre_ef, 'norma_metodo': norma_ef,
                'operador': op, 'valor': numero, 'display': display,
                'banda': not bool(nombre),
            })
    return filas


# ── Cotejo contra lo sembrado ────────────────────────────────────────────────

def clave_fiqui_legado(cuadro):
    """El cuadro sembrado → (fluido, kv) del Excel EN SERVICIO/USADO."""
    cond = cuadro['condicion']
    aceites = cond.get('oil_type_code', [])
    equipos = cond.get('equipment_type_code', [])
    desde, hasta = cond.get('voltage_from'), cond.get('voltage_to')
    if 'conmutador' in equipos:
        kv = '≤69' if hasta == 69 else '>69'
        return ('conmutador', kv)
    if 'mineral' in aceites:
        if hasta == 69 and desde is None:
            return ('mineral', '≤69')
        if desde == 69 and hasta == 230:
            return ('mineral', '>69-<230')
        if desde == 230:
            return ('mineral', '≥230')
    if 'silicona' in aceites:
        return ('silicona', None)
    if 'ester_vegetal' in aceites:
        # Los cortes sembrados son 72.5/170 (del código); el Excel usa 69/230.
        if hasta in (69, 72.5) and desde is None:
            return ('vegetal', '≤69')
        if desde in (69, 72.5):
            if hasta in (170, 230):
                return ('vegetal', '>69-230')
            return None
        if desde in (170, 230) and hasta is None:
            return ('vegetal', '≥230')
    return None


def cotejar(filas_excel, legado):
    # Índice del Excel: solo el estado que el sistema aplica (en servicio /
    # usado / conmutador en servicio), que es lo comparable con lo sembrado.
    idx_fq = {}
    for f in filas_excel:
        if f['hoja'] != 'FQ':
            continue
        comparable = (
            (f['fluido'] == 'mineral' and f['estado'] == 'en_servicio')
            or (f['fluido'] == 'conmutador' and f['estado'] == 'en_servicio')
            or (f['fluido'] in ('silicona', 'midel', 'vegetal') and f['estado'] == 'usado')
        )
        if comparable:
            idx_fq[(f['fluido'], f['kv'], f['analito'])] = f

    idx_cr = {}
    for f in filas_excel:
        if f['hoja'] != 'CR':
            continue
        idx_cr[(f['fluido'], f.get('variante'), f['equipo'], f['analito'])] = f

    coinciden, difieren, solo_codigo = [], [], []
    usadas = set()

    for cuadro in legado['cuadros']:
        etiqueta = cuadro['label']
        if cuadro['grupo'] == 'fiquis':
            clave_base = clave_fiqui_legado(cuadro)
            for lim in cuadro['limites']:
                analito = lim['analyte']
                excel = idx_fq.get((clave_base[0], clave_base[1], analito)) if clave_base else None
                comparar(etiqueta, analito, lim, excel, coinciden, difieren, solo_codigo, usadas)
        elif cuadro['grupo'] == 'cromas':
            cond = cuadro['condicion']
            aceites = cond.get('oil_type_code', [])
            equipos = cond.get('equipment_type_code', [])
            if 'conmutador' in equipos:
                clave = ('conmutador', None, None)
            elif 'mineral' in aceites:
                clave = ('mineral', None, equipos[0] if equipos else None)
            elif 'silicona' in aceites:
                clave = ('silicona', None, None)
            elif 'ester_sintetico' in aceites:
                clave = ('midel', None, None)
            elif 'ester_vegetal' in aceites:
                clave = ('vegetal', 'soya_fr3', None)
            elif etiqueta == 'Vegetal girasol':
                clave = ('vegetal', 'girasol', None)
            else:
                clave = None   # Reactor: el Excel no lo tiene
            for lim in cuadro['limites']:
                excel = idx_cr.get((clave[0], clave[1], clave[2], lim['analyte'])) if clave else None
                comparar(etiqueta, lim['analyte'], lim, excel, coinciden, difieren, solo_codigo, usadas)

    solo_excel = [f for f in filas_excel if id(f) not in usadas and limpiar(f['display']) not in ('', '-')]
    return coinciden, difieren, solo_excel, solo_codigo


def comparar(etiqueta, analito, lim, excel, coinciden, difieren, solo_codigo, usadas):
    sembrado_num = lim.get('value')
    sembrado_disp = limpiar(lim.get('display', ''))
    if excel is not None:
        usadas.add(id(excel))
    if excel is None:
        if sembrado_disp not in ('', '-'):
            solo_codigo.append({'cuadro': etiqueta, 'analito': analito, 'sembrado': sembrado_disp})
        return
    fila = {
        'cuadro': etiqueta, 'analito': analito, 'celda': f"{excel['hoja']}!{excel['celda']}",
        'excel': excel['display'], 'sembrado': sembrado_disp or '-',
    }
    ex_num = excel['valor']
    if sembrado_num is not None and ex_num is not None:
        (coinciden if float(sembrado_num) == float(ex_num) else difieren).append(fila)
    elif sembrado_disp in ('', '-') and limpiar(excel['display']) in ('', '-'):
        coinciden.append(fila)
    elif sin_acentos(sembrado_disp).lower() == sin_acentos(limpiar(excel['display'])).lower():
        coinciden.append(fila)
    elif sembrado_disp in ('', '-') or limpiar(excel['display']) in ('', '-'):
        difieren.append(fila)
    else:
        difieren.append(fila)


def tabla(filas, columnas):
    if not filas:
        return '_(ninguno)_\n'
    out = '| ' + ' | '.join(c.title() for c in columnas) + ' |\n'
    out += '|' + '---|' * len(columnas) + '\n'
    for f in filas:
        out += '| ' + ' | '.join(str(f.get(c, '')) for c in columnas) + ' |\n'
    return out


def main():
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    filas = extraer_fq(wb) + extraer_cr(wb) + extraer_otros(wb)
    legado = json.load(open(LEGADO))

    coinciden, difieren, solo_excel, solo_codigo = cotejar(filas, legado)

    datos = {
        '_meta': {
            'que_es': 'La matriz COMPLETA de VALORES_DE_ORIENTACION.xlsx (fuente primaria del laboratorio), extraída por máquina.',
            'extraido_de': 'docs/origen-ruby/fuentes-laboratorio/VALORES_DE_ORIENTACION.xlsx (hojas FQ, CR, OTROS)',
            'script': 'docs/origen-ruby/fuentes-laboratorio/etl_valores_orientacion.py',
            'nota': 'El laboratorio usó en el sistema viejo un SUBCONJUNTO de esta matriz. Este archivo es el universo; qué rige lo decide el laboratorio (ver COTEJO-EXCEL-LIMITES.md).',
        },
        'limites': filas,
    }
    SALIDA_JSON.write_text(json.dumps(datos, ensure_ascii=False, indent=2) + '\n', encoding='utf-8')

    con_valor = [f for f in filas if limpiar(f['display']) not in ('', '-')]
    md = f"""# Cotejo: el Excel del laboratorio contra lo sembrado

> Generado por `etl_valores_orientacion.py` — no editar a mano; corregir el
> script y regenerar. El Excel es la FUENTE PRIMARIA; lo sembrado salió del
> código Ruby. **Este cotejo no cambia ningún número**: es el informe con el
> que el laboratorio decide.

- Celdas con criterio en el Excel: **{len(con_valor)}** (de {len(filas)} leídas).
- Límites sembrados comparables encontrados en el Excel: **{len(coinciden) + len(difieren)}**.

| Categoría | Cantidad | Qué significa |
|---|---|---|
| **Coinciden** | {len(coinciden)} | El código viejo implementó fiel el Excel: quedan confirmados. |
| **Difieren** | {len(difieren)} | El Excel dice una cosa y lo sembrado (= el código viejo) otra. Decide el laboratorio cuál manda. |
| **Solo en el Excel** | {len(solo_excel)} | Criterios que el sistema viejo nunca implementó (estados nuevo/antes de energizar/tratado, rigidez IEC 60156, hoja OTROS…). El laboratorio elige cuáles activar. |
| **Solo en el código** | {len(solo_codigo)} | Sembrado sin respaldo en el Excel (ej. los cuadros de Reactor). |

## 1. Difieren — A DECIDIR

{tabla(difieren, ['cuadro', 'analito', 'excel', 'sembrado', 'celda'])}

## 2. Solo en el código — sin respaldo en el Excel

{tabla(solo_codigo, ['cuadro', 'analito', 'sembrado'])}

## 3. Solo en el Excel — nunca implementados (elegir cuáles activar)

{tabla([{'celda': f"{f['hoja']}!{f['celda']}", 'fluido': f['fluido'], 'estado': f.get('estado') or '-', 'kv': f.get('kv') or '-', 'equipo': f.get('equipo') or f.get('variante') or '-', 'analito': f.get('analito') or f.get('parametro'), 'criterio': f['display']} for f in solo_excel], ['celda', 'fluido', 'estado', 'kv', 'equipo', 'analito', 'criterio'])}

## 4. Coinciden — confirmados contra la fuente primaria

{tabla(coinciden, ['cuadro', 'analito', 'excel', 'sembrado', 'celda'])}
"""
    SALIDA_COTEJO.write_text(md, encoding='utf-8')
    print(f'filas Excel: {len(filas)} (con criterio: {len(con_valor)})')
    print(f'coinciden: {len(coinciden)} | difieren: {len(difieren)} | solo Excel: {len(solo_excel)} | solo código: {len(solo_codigo)}')


if __name__ == '__main__':
    sys.exit(main())
